import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill

# Funciones de lectura y comparación
def read_bom(file_path):
    bom_dict = {}
    with open(file_path, newline='', encoding='utf-8-sig') as csvfile:
        reader = csv.DictReader(csvfile)
        ref_col = reader.fieldnames[0]
        for row in reader:
            ref = row[ref_col].strip()
            bom_dict[ref] = {k: v for k, v in row.items() if k != ref_col}
    return bom_dict

def compare_boms(bom_v0, bom_v1):
    modified = {}
    added = {}
    removed = {}

    for ref, data in bom_v0.items():
        if ref in bom_v1:
            changes = {}
            for key in data.keys():
                if data[key] != bom_v1[ref].get(key, ''):
                    changes[key] = {'V0': data[key], 'V1': bom_v1[ref].get(key, '')}
            if changes:
                modified[ref] = changes
        else:
            removed[ref] = data

    for ref, data in bom_v1.items():
        if ref not in bom_v0:
            added[ref] = data

    return modified, added, removed

# Generar XLSX
def write_changes_xlsx(modified, added, removed, filename='BOM_Comparison.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Comparison"

    # Encabezado
    headers = ['Ref Des', 'Change Type', 'Field', 'V0', 'V1']
    ws.append(headers)

    # Alineación centrada para combinar celdas
    center_align = Alignment(vertical='center', horizontal='center')

    # Styles: fills and border
    fill_modified = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    fill_added = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_removed = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    thick = Side(border_style='thick', color='000000')
    none_side = Side(border_style=None)

    # Modificados
    for ref, changes in modified.items():
        fields = list(changes.keys())
        start_row = ws.max_row + 1
        for field in fields:
            ws.append([ref, 'Modified', field, changes[field]['V0'], changes[field]['V1']])
        end_row = ws.max_row
        # Combinar celdas de Ref Des
        if len(fields) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = center_align
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.cell(row=start_row, column=2).alignment = center_align
        # Apply fill and thick outer border for the block
        for r in range(start_row, end_row + 1):
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_modified
                top = thick if r == start_row else none_side
                bottom = thick if r == end_row else none_side
                left = thick if c == 1 else none_side
                right = thick if c == 5 else none_side
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Añadidos
    for ref, data in added.items():
        fields = list(data.keys())
        start_row = ws.max_row + 1
        for field in fields:
            ws.append([ref, 'Added', field, '', data.get(field, '')])
        end_row = ws.max_row
        if len(fields) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = center_align
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.cell(row=start_row, column=2).alignment = center_align
        # Apply fill and thick outer border for the block
        for r in range(start_row, end_row + 1):
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_added
                top = thick if r == start_row else none_side
                bottom = thick if r == end_row else none_side
                left = thick if c == 1 else none_side
                right = thick if c == 5 else none_side
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Eliminados
    for ref, data in removed.items():
        fields = list(data.keys())
        start_row = ws.max_row + 1
        for field in fields:
            ws.append([ref, 'Removed', field, data.get(field, ''), ''])
        end_row = ws.max_row
        if len(fields) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = center_align
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.cell(row=start_row, column=2).alignment = center_align
        # Apply fill and thick outer border for the block
        for r in range(start_row, end_row + 1):
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_removed
                top = thick if r == start_row else none_side
                bottom = thick if r == end_row else none_side
                left = thick if c == 1 else none_side
                right = thick if c == 5 else none_side
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Ajustar ancho de columnas
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filename)
    print(f"BOM comparada y guardada en '{filename}'")

# -------------------------
# Ejecutar comparador y generar XLSX
bom_v0 = read_bom('V0.csv')
bom_v1 = read_bom('V1.csv')

modified, added, removed = compare_boms(bom_v0, bom_v1)
write_changes_xlsx(modified, added, removed)
